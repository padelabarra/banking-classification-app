"""
Aggregations for dashboard charts.
"""
import pandas as pd


def monthly_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Monthly Revenue vs Expense totals.
    Returns: year, month, period, revenue, expense, net
    """
    if df.empty:
        return pd.DataFrame(columns=["year", "month", "period", "revenue", "expense", "net"])

    df = df.copy()
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0)

    g = (
        df.groupby(["year", "month", "activity"])["amount"]
        .sum()
        .unstack(fill_value=0)
        .reset_index()
    )

    if "Revenue" not in g.columns:
        g["Revenue"] = 0.0
    if "Expense" not in g.columns:
        g["Expense"] = 0.0

    g["expense"] = g["Expense"].abs()
    g["revenue"] = g["Revenue"]
    g["net"] = g["revenue"] - g["expense"]
    g["period"] = g.apply(lambda r: f"{int(r['year'])}-{int(r['month']):02d}", axis=1)

    return g[["year", "month", "period", "revenue", "expense", "net"]].sort_values(
        ["year", "month"]
    )


def category_breakdown(df: pd.DataFrame, activity: str = "Expense") -> pd.DataFrame:
    """
    Total by category for a given activity filter.
    Returns: category, total (absolute), count
    """
    if df.empty:
        return pd.DataFrame(columns=["category", "total", "count"])

    filtered = df[df["activity"] == activity].copy()
    filtered["amount"] = pd.to_numeric(filtered["amount"], errors="coerce").fillna(0)

    result = (
        filtered.groupby("category")["amount"]
        .agg(total="sum", count="count")
        .reset_index()
    )
    result["total"] = result["total"].abs()
    return result.sort_values("total", ascending=False)


def category_trend(df: pd.DataFrame, top_n: int = 8) -> pd.DataFrame:
    """
    Monthly totals per top-N categories over time.
    Returns: period, category, total
    """
    if df.empty:
        return pd.DataFrame(columns=["period", "category", "total"])

    expenses = df[df["activity"] == "Expense"].copy()
    expenses["amount"] = pd.to_numeric(expenses["amount"], errors="coerce").fillna(0).abs()
    expenses["period"] = expenses.apply(
        lambda r: f"{int(r['year'])}-{int(r['month']):02d}", axis=1
    )

    # Top N categories by overall spend
    top_cats = (
        expenses.groupby("category")["amount"].sum().nlargest(top_n).index.tolist()
    )
    filtered = expenses[expenses["category"].isin(top_cats)]

    result = (
        filtered.groupby(["period", "category"])["amount"].sum().reset_index(name="total")
    )
    return result.sort_values("period")


def source_breakdown(df: pd.DataFrame) -> pd.DataFrame:
    """Expense totals by source."""
    if df.empty:
        return pd.DataFrame(columns=["source", "total", "count"])

    expenses = df[df["activity"] == "Expense"].copy()
    expenses["amount"] = pd.to_numeric(expenses["amount"], errors="coerce").fillna(0).abs()

    return (
        expenses.groupby("source")["amount"]
        .agg(total="sum", count="count")
        .reset_index()
        .sort_values("total", ascending=False)
    )


def cumulative_net(df: pd.DataFrame) -> pd.DataFrame:
    """Monthly net savings with running cumulative."""
    summary = monthly_summary(df)
    if summary.empty:
        return summary
    summary = summary.sort_values(["year", "month"])
    summary["cumulative_net"] = summary["net"].cumsum()
    return summary


import pathlib as _pathlib

_EXCEL_PATH = _pathlib.Path(__file__).parent.parent / "202406_Presupuesto_MBA.xlsx"

_TRACKER_RECURRING_ROWS = {
    7:  "Car Gasoline / Transportation",
    8:  "Car Insurance",
    9:  "Dinning&Activities",
    10: "Groceries",
    11: "Miscellaneous",
    12: "Online Shopping",
    13: "Parking",
    14: "Housing",
    15: "Living Expenses",
}

_TRACKER_BIG_TICKET_ROWS = {
    2: "Car",
    3: "Pack",
    4: "Trips",
    5: "Tuition",
}


def load_budget_defaults() -> dict[str, float]:
    """
    Load monthly budget defaults from the Excel Tracker sheet.

    For recurring categories (rows 7-15): col C is the historical monthly average.
    For big-ticket categories (rows 2-5): col C is the total sum, divided by 12.
    Returns an empty dict if the Excel file cannot be read.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(str(_EXCEL_PATH), data_only=True)
        ws = wb["Tracker"]
    except Exception:
        return {}

    budgets: dict[str, float] = {}

    for row_num, category in _TRACKER_RECURRING_ROWS.items():
        val = ws.cell(row=row_num, column=3).value  # col C
        budgets[category] = abs(float(val)) if val is not None else 0.0

    for row_num, category in _TRACKER_BIG_TICKET_ROWS.items():
        val = ws.cell(row=row_num, column=3).value  # col C
        budgets[category] = abs(float(val)) / 12.0 if val is not None else 0.0

    return budgets


def monthly_actuals_by_category(df: pd.DataFrame, year: int, month: int) -> dict[str, float]:
    """
    Sum of absolute expense amounts per category for a given year/month.

    Returns: {category: total_spent (positive float)}
    """
    mask = (
        (df["activity"] == "Expense") &
        (pd.to_numeric(df["year"], errors="coerce") == year) &
        (pd.to_numeric(df["month"], errors="coerce") == month)
    )
    sub = df[mask].copy()
    if sub.empty:
        return {}
    sub["amount"] = pd.to_numeric(sub["amount"], errors="coerce").fillna(0).abs()
    return sub.groupby("category")["amount"].sum().to_dict()
